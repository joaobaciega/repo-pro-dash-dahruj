"""
Fase 5 — App integrado (app.py)

App único com telas navegáveis pelo menu lateral:
  - Dashboard: acompanhamento trimestral móvel (lê da view vw_base_tidy no MySQL).
  - Relatório Por Gerente / Por Consultor: rankings do mês.
  - Lançamento (restrito): o gestor informa Passagens e Refis por consultor.
    Não aparece no menu; só entra depois da senha pedida pelo cadeado no fim da
    barra lateral (ver bloco ACESSO AO LANÇAMENTO).

Tudo sai do banco: ao salvar um lançamento, o cache de leitura é limpo e o
dashboard reflete a mudança na hora. Não usa mais Excel.

Como rodar (na pasta do projeto, com .streamlit/secrets.toml configurado):
    streamlit run app.py
"""

import datetime as dt
import hmac
import io
import time
from pathlib import Path

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

import db

st.set_page_config(page_title="Dahruj — Refis para Palhetas",
                   page_icon="📊", layout="wide")

N_MESES = 3
# Identidade visual: PRETO / LARANJA / BRANCO (proporção 60/30/10).
ORANGE = "#EB5E33"          # laranja da marca — cor de acento forte (30%)
ORANGE_ESC = "#B5451F"      # laranja escuro (variação para séries)
PRETO = "#0D0D0D"           # fundo dominante (60%)
BRANCO = "#F5F5F5"          # texto/acento claro (10%)
ASSETS = Path(__file__).resolve().parent / "assets"
# Paleta dos gráficos: laranja em primeiro plano, com tons e branco/cinzas para
# diferenciar séries mantendo a leitura limpa sobre fundo preto.
PALETTE = ["#EB5E33", "#F5A623", "#FF8A5B", "#F58220", "#FFFFFF",
           "#C0C0C0", "#B5451F", "#8A8A8A", "#FFB07C", "#E0E0E0"]
MESES_PT = ["", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
            "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]

# --------------------------- Acesso ao Lançamento ---------------------------
# Dashboard e relatórios são abertos; o Lançamento fica oculto atrás do cadeado
# no fim da barra lateral e só entra no menu depois da senha.
PAG_DASHBOARD = "Dashboard"
PAG_VERBAS = "Verbas"
PAG_GERENTE = "Relatório Por Gerente"
PAG_CONSULTOR = "Relatório Por Consultor"
PAG_LANCAMENTO = "📝 Lançamento"
PAGINAS_PUBLICAS = [PAG_DASHBOARD, PAG_GERENTE, PAG_CONSULTOR, PAG_VERBAS]

SENHA_LANCAMENTO = "Dahruj00$"   # pode ser trocada em secrets: [acesso] senha_lancamento
_TTL_LANCAMENTO = 30 * 60        # segundos de inatividade até travar de novo
_MAX_TENTATIVAS = 3
_ESPERA_BLOQUEIO = 60            # segundos de espera após esgotar as tentativas


# --------------------------- Identidade visual ---------------------------
def _injetar_css():
    """Aplica a identidade PRETO/LARANJA/BRANCO (60/30/10) sobre o tema escuro."""
    st.markdown(f"""
        <style>
        /* Títulos e cabeçalhos em laranja (acento de marca) */
        h1, h2, h3 {{ color: {ORANGE} !important; }}
        /* Cartões de KPI: fundo escuro com acento laranja à esquerda */
        div[data-testid="stMetric"] {{
            background: #1A1A1A; border-left: 5px solid {ORANGE};
            border-radius: 8px; padding: 14px 16px; }}
        div[data-testid="stMetricValue"] {{ color: {BRANCO}; }}
        /* Verba gerada: 2px abaixo do padrão do Streamlit (2.25rem = 36px) para
           os centavos de "Total gerado" não estourarem em telas estreitas. */
        .st-key-verba_gerada div[data-testid="stMetricValue"] {{ font-size: 34px; }}
        /* Sidebar mais escura, com borda laranja sutil */
        section[data-testid="stSidebar"] {{
            background: #0A0A0A; border-right: 1px solid rgba(235,94,51,.35); }}
        /* Divisores em laranja translúcido */
        hr {{ border-color: rgba(235,94,51,.45) !important; }}
        /* Item de menu (radio) selecionado destacado em laranja */
        section[data-testid="stSidebar"] label[data-baseweb="radio"]:has(input:checked) {{
            color: {ORANGE}; font-weight: 700; }}
        </style>
    """, unsafe_allow_html=True)


def _achar_logo():
    for ext in ("png", "jpg", "jpeg", "webp", "gif"):
        p = ASSETS / f"logo.{ext}"
        if p.exists():
            return p
    return None


def mostrar_logo():
    """Exibe a logo grande no topo da barra lateral — canto superior esquerdo,
    presente em todas as páginas. Sem arquivo, cai para o texto 'DAHRUJ'.
    """
    logo = _achar_logo()
    if logo:
        st.sidebar.image(str(logo), width="stretch")
    else:
        st.sidebar.markdown(
            f'<div style="font-family:\'Arial Black\',Arial,sans-serif;'
            f'font-weight:900;font-size:30px;letter-spacing:2px;color:{ORANGE};'
            f'border:3px solid {ORANGE};border-radius:8px;padding:2px 12px;'
            f'display:inline-block;margin-bottom:6px;">DAHRUJ</div>',
            unsafe_allow_html=True)


# --------------------------- Formatação BR ---------------------------
def fmt_money(v):
    if v is None or pd.isna(v):
        return "—"
    return ("R$ {:,.2f}".format(v)).replace(",", "X").replace(".", ",").replace("X", ".")


def fmt_int(v):
    if v is None or pd.isna(v):
        return "—"
    return "{:,.0f}".format(v).replace(",", ".")


def fmt_pct(v):
    if v is None or pd.isna(v):
        return "—"
    return ("{:.1f}%".format(v * 100)).replace(".", ",")


def delta_str(v, kind):
    if v is None or pd.isna(v):
        return None
    sign = "+" if v >= 0 else "-"
    a = abs(v)
    if kind == "money":
        body = fmt_money(a)
    elif kind == "int":
        body = fmt_int(a)
    else:  # pp (pontos percentuais)
        body = "{:.1f}".format(a * 100).replace(".", ",") + " p.p."
    return f"{sign}{body}"


def mlabel(d):
    d = pd.Timestamp(d)
    return f"{MESES_PT[d.month][:3]}/{d.year}"


# --------------------------- Helpers de dados ---------------------------
def last_n_months(df, n=N_MESES, mes_fim=None):
    """Janela dos n meses que TERMINAM em `mes_fim` (trimestre móvel). Sem
    `mes_fim`, usa o mês mais recente da base."""
    meses = sorted(df["mes"].unique())
    if mes_fim is not None and mes_fim in meses:
        idx = meses.index(mes_fim)
        keep = meses[max(0, idx - n + 1): idx + 1]
    else:
        keep = meses[-n:]
    return df[df["mes"].isin(keep)].copy(), keep


def agg_by(df, group):
    """Agrega os KPIs. O aproveitamento é SEMPRE recalculado como
    soma(refil_diant) / soma(passagens) — nunca média de percentuais. Para a
    conversão, conta apenas refis de linhas que também têm passagens informadas,
    evitando inflar a taxa com linhas incompletas. O aproveitamento traseiro
    (refil_tras / passagens) é calculado com o mesmo critério."""
    df = df.copy()
    df["_refil_conv"] = df["refil_diant"].where(df["passagens"].notna())
    df["_refil_conv_t"] = df["refil_tras"].where(df["passagens"].notna())
    g = df.groupby(group, as_index=False).agg(
        passagens=("passagens", "sum"),
        refil_diant=("refil_diant", "sum"),
        refil_tras=("refil_tras", "sum"),
        total_diant=("total_diant", "sum"),
        total_tras=("total_tras", "sum"),
        total_geral=("total_geral", "sum"),
        _refil_conv=("_refil_conv", "sum"),
        _refil_conv_t=("_refil_conv_t", "sum"),
    )
    g["aproveitamento"] = (g["_refil_conv"] / g["passagens"]).where(g["passagens"] > 0)
    g["aproveitamento_tras"] = (g["_refil_conv_t"] / g["passagens"]).where(g["passagens"] > 0)
    return g.drop(columns=["_refil_conv", "_refil_conv_t"])


def gerar_excel_ranking(f, periodo, filtros_txt):
    """Gera, em memória, um relatório Excel formatado do Ranking da seleção atual.
    Aba 1: ranking por consultor no trimestre. Aba 2: resumo mensal."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    NAVY, WHITE = "FF1F3864", "FFFFFFFF"
    thin = Side(style="thin", color="FFD9D9D9")
    BD = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _v(x):  # NaN -> None (célula vazia em vez de "nan")
        return None if (isinstance(x, float) and pd.isna(x)) else x

    def _i(x):
        return int(x) if pd.notna(x) else 0

    def cab(ws, headers, row):
        for j, h in enumerate(headers, 1):
            c = ws.cell(row, j, h)
            c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
            c.fill = PatternFill("solid", fgColor=NAVY)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = BD

    wb = Workbook()

    # ---- Aba 1: Ranking por consultor (trimestre) ----
    ws = wb.active
    ws.title = "Ranking Trimestre"
    ws.cell(1, 1, "Ranking Vendas Refis para Palhetas — Dahruj").font = \
        Font(name="Arial", bold=True, size=14, color=NAVY)
    for r, txt in ((2, f"Período: {periodo}"), (3, f"Filtros: {filtros_txt}"),
                   (4, f"Gerado em: {dt.datetime.now().strftime('%d/%m/%Y %H:%M')}")):
        ws.cell(r, 1, txt).font = Font(name="Arial", size=9, color="FF666666")

    rk = agg_by(f, ["consultor", "unidade", "marca"]).sort_values("total_geral", ascending=False)
    overall = agg_by(f.assign(_g=1), "_g").iloc[0]
    head = ["#", "Consultor", "Unidade", "Marca", "Passagens", "Refil Diant.",
            "Aprov. Diant.", "Refil Tras.", "Aprov. Tras.", "Faturamento (R$)"]
    H = 6
    cab(ws, head, H)
    for i, (_, row) in enumerate(rk.iterrows()):
        r = H + 1 + i
        vals = [i + 1, row["consultor"], row["unidade"], row["marca"],
                row["passagens"], row["refil_diant"], row["aproveitamento"],
                row["refil_tras"], row["aproveitamento_tras"], row["total_geral"]]
        for j, v in enumerate(vals, 1):
            c = ws.cell(r, j, _v(v)); c.font = Font(name="Arial", size=10); c.border = BD
            if j in (1, 5, 6, 7, 8, 9):
                c.alignment = Alignment(horizontal="center")
            if j in (5, 6, 8):
                c.number_format = "#,##0"
            elif j in (7, 9):
                c.number_format = "0.0%"
            elif j == 10:
                c.number_format = "R$ #,##0.00"
        if i % 2 == 1:
            for j in range(1, len(head) + 1):
                ws.cell(r, j).fill = PatternFill("solid", fgColor="FFF4F6FA")
    tr = H + 1 + len(rk)
    ws.cell(tr, 4, "TOTAL").font = Font(name="Arial", bold=True, size=10)
    for j, v, fmt in ((5, _i(overall["passagens"]), "#,##0"),
                      (6, _i(overall["refil_diant"]), "#,##0"),
                      (7, _v(overall["aproveitamento"]), "0.0%"),
                      (8, _i(overall["refil_tras"]), "#,##0"),
                      (9, _v(overall["aproveitamento_tras"]), "0.0%"),
                      (10, _v(overall["total_geral"]), "R$ #,##0.00")):
        c = ws.cell(tr, j, v); c.font = Font(name="Arial", bold=True, size=10)
        c.border = BD; c.number_format = fmt
        if j in (5, 6, 7, 8, 9):
            c.alignment = Alignment(horizontal="center")
    for j, w in enumerate([5, 28, 16, 9, 11, 12, 12, 11, 12, 15], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = f"A{H + 1}"

    # ---- Aba 2: Resumo mensal ----
    ws2 = wb.create_sheet("Resumo Mensal")
    mensal = agg_by(f, "mes").sort_values("mes")
    mensal["lbl"] = mensal["mes"].apply(mlabel)
    head2 = ["Mês", "Passagens", "Refil Diant.", "Aprov. Diant.", "Refil Tras.",
             "Aprov. Tras.", "Fat. Diant. (R$)", "Fat. Tras. (R$)", "Fat. Total (R$)"]
    cab(ws2, head2, 1)
    for i, (_, row) in enumerate(mensal.iterrows()):
        r = 2 + i
        vals = [row["lbl"], row["passagens"], row["refil_diant"], row["aproveitamento"],
                row["refil_tras"], row["aproveitamento_tras"], row["total_diant"],
                row["total_tras"], row["total_geral"]]
        for j, v in enumerate(vals, 1):
            c = ws2.cell(r, j, _v(v)); c.font = Font(name="Arial", size=10); c.border = BD
            if j in (2, 3, 5):
                c.number_format = "#,##0"; c.alignment = Alignment(horizontal="center")
            elif j in (4, 6):
                c.number_format = "0.0%"; c.alignment = Alignment(horizontal="center")
            elif j in (7, 8, 9):
                c.number_format = "R$ #,##0.00"
    for j, w in enumerate([12, 11, 12, 12, 11, 12, 14, 14, 14], 1):
        ws2.column_dimensions[get_column_letter(j)].width = w
    ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================ PÁGINA: DASHBOARD ============================
def pagina_dashboard():
    st.title("📊 Análise de Vendas — Dahruj")
    try:
        df = db.ler_base_tidy()
    except Exception:
        st.error("Não foi possível conectar ao banco de dados no momento. "
                 "Verifique se o MySQL está ativo e tente novamente.")
        return

    if df.empty:
        st.info("Ainda não há lançamentos no banco. Procure o gestor responsável "
                "pelo lançamento dos dados.")
        return

    # ---- Seleção do mês de referência (menu suspenso) ----
    meses_disp = sorted(df["mes"].dropna().unique())
    lbl_to_mes = {mlabel(m): m for m in meses_disp}
    labels_mes = [mlabel(m) for m in reversed(meses_disp)]  # mais recente primeiro
    csel, _ = st.columns([2, 3])
    mes_sel_lbl = csel.selectbox(
        "📅 Mês de referência", labels_mes, index=0,
        help="Escolha o mês para ver os KPIs. Os gráficos mostram o trimestre "
             "móvel encerrado no mês selecionado.")
    mes_sel = lbl_to_mes[mes_sel_lbl]

    win, keep = last_n_months(df, mes_fim=mes_sel)
    ordem_meses = [mlabel(m) for m in sorted(keep)]
    periodo = " · ".join(ordem_meses)
    st.caption(f"Acompanhamento trimestral móvel · {periodo}")

    # ---- Filtros (sidebar) ----
    st.sidebar.header("Filtros")
    st.sidebar.caption(f"Trimestre vigente: **{periodo}**")
    marcas = sorted(win["marca"].dropna().unique())
    sel_marcas = st.sidebar.multiselect("Marca", marcas, default=[])

    base_uni = win[win["marca"].isin(sel_marcas)] if sel_marcas else win
    unidades = sorted(base_uni["unidade"].dropna().unique())
    sel_unidades = st.sidebar.multiselect("Unidade", unidades, default=[])

    base_cons = win.copy()
    if sel_marcas:
        base_cons = base_cons[base_cons["marca"].isin(sel_marcas)]
    if sel_unidades:
        base_cons = base_cons[base_cons["unidade"].isin(sel_unidades)]
    consultores = sorted(base_cons["consultor"].dropna().unique())
    sel_cons = st.sidebar.multiselect("Consultor", consultores, default=[])
    st.sidebar.divider()
    st.sidebar.caption("Sem consultor selecionado, mostra o agregado do filtro. "
                       "Selecionando consultores, cada um vira uma linha.")

    f = win.copy()
    if sel_marcas:
        f = f[f["marca"].isin(sel_marcas)]
    if sel_unidades:
        f = f[f["unidade"].isin(sel_unidades)]
    if sel_cons:
        f = f[f["consultor"].isin(sel_cons)]

    if f.empty:
        st.warning("Nenhum dado para os filtros selecionados.")
        return

    # ---- Exportar relatório (reflete os filtros atuais) ----
    partes = []
    if sel_marcas:
        partes.append("Marcas: " + ", ".join(sel_marcas))
    if sel_unidades:
        partes.append("Unidades: " + ", ".join(sel_unidades))
    if sel_cons:
        partes.append("Consultores: " + ", ".join(sel_cons))
    filtros_txt = " · ".join(partes) if partes else "Todos"
    _, colexp = st.columns([3, 1])
    colexp.download_button(
        "📥 Exportar relatório (Excel)",
        data=gerar_excel_ranking(f, periodo, filtros_txt),
        file_name=f"Ranking_Dahruj_{dt.date.today().isoformat()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    # ---- KPI cards ----
    mensal = agg_by(f, "mes").sort_values("mes")
    mensal["mes_lbl"] = mensal["mes"].apply(mlabel)
    atual = mensal.iloc[-1]
    ant = mensal.iloc[-2] if len(mensal) > 1 else None

    def d(col, kind):
        return delta_str(atual[col] - ant[col], kind) if ant is not None else None

    r1 = st.columns(3)
    r1[0].metric(f"Passagens ({mes_sel_lbl})", fmt_int(atual["passagens"]), d("passagens", "int"))
    r1[1].metric(f"Nº Refil Diant. ({mes_sel_lbl})", fmt_int(atual["refil_diant"]),
                 d("refil_diant", "int"))
    r1[2].metric(f"Aproveitamento Diant. ({mes_sel_lbl})", fmt_pct(atual["aproveitamento"]),
                 d("aproveitamento", "pp"))
    r2 = st.columns(3)
    r2[0].metric(f"Faturamento total ({mes_sel_lbl})", fmt_money(atual["total_geral"]),
                 d("total_geral", "money"))
    r2[1].metric(f"Nº Refil Tras. ({mes_sel_lbl})", fmt_int(atual["refil_tras"]),
                 d("refil_tras", "int"))
    r2[2].metric(f"Aproveitamento Tras. ({mes_sel_lbl})", fmt_pct(atual["aproveitamento_tras"]),
                 d("aproveitamento_tras", "pp"))
    st.caption("Variação comparada ao mês imediatamente anterior dentro da janela.")
    st.divider()

    # ---- Evolução ----
    st.subheader("Evolução no trimestre")
    by_cons = bool(sel_cons)
    if by_cons:
        evo = agg_by(f, ["mes", "consultor"]).sort_values("mes")
        evo["mes_lbl"] = evo["mes"].apply(mlabel)
        color = "consultor"
    else:
        evo = mensal.copy()
        evo["consultor"] = "Agregado"
        color = None

    def line_fig(dfp, ycol, titulo, fmt):
        fig = px.line(dfp, x="mes_lbl", y=ycol, color=color, markers=True,
                      color_discrete_sequence=PALETTE,
                      category_orders={"mes_lbl": ordem_meses})
        fig.update_traces(hovertemplate="%{x}<br>" + fmt + "<extra>%{fullData.name}</extra>")
        fig.update_layout(title=titulo, template="plotly_dark", height=300,
                          margin=dict(l=10, r=10, t=50, b=10),
                          xaxis_title=None, yaxis_title=None,
                          paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                          font_color=BRANCO,
                          legend=dict(orientation="h", y=-0.2, title=None))
        # Piso do eixo Y por métrica (em vez de começar no zero): dá zoom o
        # suficiente para ver a variação, sem exagerar as quedas. O piso só vale
        # quando os dados estão acima dele — no modo por consultor, com valores
        # menores, o eixo volta a se ajustar sozinho (não corta as linhas).
        pisos = {"passagens": 1500, "refil_diant": 350,
                 "aproveitamento": 0.10, "total_geral": 80000}
        piso = pisos.get(ycol)
        ymax = dfp[ycol].max()
        if piso is not None and pd.notna(ymax) and ymax > piso:
            fig.update_yaxes(range=[piso, ymax * 1.05])
        else:
            fig.update_yaxes(rangemode="tozero")
        if ycol == "aproveitamento":
            fig.update_yaxes(tickformat=".0%")
        return fig

    r1c1, r1c2 = st.columns(2)
    r1c1.plotly_chart(line_fig(evo, "passagens", "Passagens", "%{y:,.0f}"),
                      use_container_width=True)
    r1c2.plotly_chart(line_fig(evo, "refil_diant", "Nº Refil Dianteiro", "%{y:,.0f}"),
                      use_container_width=True)
    r2c1, r2c2 = st.columns(2)
    r2c1.plotly_chart(line_fig(evo, "aproveitamento", "Aproveitamento", "%{y:.1%}"),
                      use_container_width=True)
    r2c2.plotly_chart(line_fig(evo, "total_geral", "Faturamento (R$)", "R$ %{y:,.2f}"),
                      use_container_width=True)
    st.divider()

    # ---- Por unidade (aproveitamento ou faturamento) ----
    ctit, cord = st.columns([3, 1])
    ord_uni = cord.radio("Ordenar por", ["Aproveitamento", "Faturamento"],
                         key="ord_unidade")
    por_fat = ord_uni == "Faturamento"
    ctit.subheader("Faturamento por unidade" if por_fat else "Aproveitamento por unidade")
    ctit.caption("Faturamento total por unidade, somado no trimestre." if por_fat
                 else "Taxa de conversão (refil dianteiro ÷ passagens) por unidade, "
                      "somada no trimestre.")
    apu = agg_by(f, "unidade")
    if por_fat:
        apu = apu.sort_values("total_geral")
        xcol = "total_geral"
        txt = apu["total_geral"].map(fmt_money)
        htmpl, xfmt = "%{y}<br>R$ %{x:,.2f}<extra></extra>", None
    else:
        apu = apu[apu["passagens"] > 0].sort_values("aproveitamento")
        xcol = "aproveitamento"
        txt = apu["aproveitamento"].map(fmt_pct)
        htmpl, xfmt = "%{y}<br>%{x:.1%}<extra></extra>", ".0%"
    if apu.empty:
        st.caption("Sem dados suficientes no período para montar este gráfico.")
    else:
        figu = px.bar(apu, x=xcol, y="unidade", orientation="h",
                      color_discrete_sequence=[ORANGE], text=txt)
        figu.update_traces(hovertemplate=htmpl, textposition="outside", cliponaxis=False)
        figu.update_layout(template="plotly_dark", height=max(320, 28 * len(apu)),
                           margin=dict(l=10, r=60, t=20, b=10),
                           xaxis_title=ord_uni, yaxis_title=None,
                           paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                           font_color=BRANCO)
        if xfmt:
            figu.update_xaxes(tickformat=xfmt)
        st.plotly_chart(figu, use_container_width=True)
    st.divider()

    # ---- Ranking ----
    st.subheader("Ranking no trimestre")
    rk = {"Faturamento (R$)": ("total_geral", "money"),
          "Passagens": ("passagens", "int"),
          "Nº Refil Dianteiro": ("refil_diant", "int"),
          "Aproveitamento": ("aproveitamento", "pct")}
    colr1, colr2 = st.columns([2, 1])
    metrica = colr1.selectbox("Ordenar por", list(rk.keys()))
    topn = colr2.slider("Quantos exibir", 5, 65, 15)
    ycol, kind = rk[metrica]
    rank = agg_by(f, "consultor").sort_values(ycol, ascending=False).head(topn).sort_values(ycol)
    if kind == "money":
        txt = rank[ycol].map(fmt_money); htmpl = "R$ %{x:,.2f}<extra></extra>"; xfmt = None
    elif kind == "pct":
        txt = rank[ycol].map(fmt_pct); htmpl = "%{x:.1%}<extra></extra>"; xfmt = ".0%"
    else:
        txt = rank[ycol].map(fmt_int); htmpl = "%{x:,.0f}<extra></extra>"; xfmt = None
    figr = px.bar(rank, x=ycol, y="consultor", orientation="h",
                  color_discrete_sequence=[ORANGE], text=txt)
    figr.update_traces(hovertemplate=htmpl, textposition="outside", cliponaxis=False)
    figr.update_layout(template="plotly_dark", height=max(340, 26 * len(rank)),
                       margin=dict(l=10, r=40, t=20, b=10),
                       xaxis_title=metrica, yaxis_title=None,
                       paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                       font_color=BRANCO)
    if xfmt:
        figr.update_xaxes(tickformat=xfmt)
    st.plotly_chart(figr, use_container_width=True)
    st.divider()

    # ---- Tabela detalhada ----
    with st.expander("Ver dados detalhados (filtro aplicado)"):
        show = f[["consultor", "unidade", "mes_label", "passagens", "refil_diant",
                  "refil_tras", "aproveitamento", "total_diant", "total_tras",
                  "total_geral"]].copy()
        show = show.rename(columns={
            "consultor": "Consultor", "unidade": "Unidade", "mes_label": "Mês",
            "passagens": "Passagens", "refil_diant": "Refil Diant.",
            "refil_tras": "Refil Tras.", "aproveitamento": "Aproveitamento",
            "total_diant": "Total Diant.", "total_tras": "Total Tras.",
            "total_geral": "Total Geral"})
        st.dataframe(show, use_container_width=True, hide_index=True,
                     column_config={
                         "Aproveitamento": st.column_config.NumberColumn(format="%.1f%%"),
                         "Total Diant.": st.column_config.NumberColumn(format="R$ %.2f"),
                         "Total Tras.": st.column_config.NumberColumn(format="R$ %.2f"),
                         "Total Geral": st.column_config.NumberColumn(format="R$ %.2f")})


# ============================== PÁGINA: VERBAS ==============================
# Esta página NÃO usa a view vw_base_tidy. Ela lê `vendas_verbas`, que vem da
# planilha "Base de dados para Dash Board.xlsx" (sell-in: pedidos faturados para
# as concessionárias, linha a linha por produto) via `importar_verbas.py`.
#
# São bases diferentes de propósito: o Dashboard mede o sell-out por consultor
# (passagens → refis), aqui medimos a verba gerada por venda. Por isso os cards
# daqui saem TODOS da mesma base — misturar as duas fontes numa linha de KPIs
# produziria números que não fecham entre si.
#
# Passagens e Aproveitamento não aparecem aqui: a base de vendas não tem
# passagens, e não há como derivá-las.
VERBAS_CATS = [("consultor", "Consultor", "total_consultor"),
               ("gerente", "Gerente", "total_gerente"),
               ("marketing", "Marketing", "total_reserva")]
OPCAO_ANO = "Ano todo"


def _resumo_verbas(df, pagos):
    """Verba gerada, paga e saldo do recorte `df`, por categoria e no total.

    Paga = só as categorias marcadas em `verbas_pagamentos`, mês a mês. Cada
    categoria tem sua própria marcação, então um mês pode ter consultor pago e
    gerente em aberto.

    Marketing nunca entra em "paga": a reserva não é paga a ninguém, é sempre
    saldo. É isso que faz Janeiro entrar inteiro no saldo — naquele mês nada foi
    pago a consultor nem a gerente, tudo virou marketing.
    """
    gerada = {cat: float(df[col].sum()) for cat, _, col in VERBAS_CATS}
    paga = {cat: 0.0 for cat, _, _ in VERBAS_CATS}

    if not df.empty:
        por_mes = df.groupby(df["data"].dt.to_period("M"))
        for periodo, g in por_mes:
            flags = pagos.get(pd.Timestamp(periodo.start_time), {})
            if flags.get("consultor"):
                paga["consultor"] += float(g["total_consultor"].sum())
            if flags.get("gerente"):
                paga["gerente"] += float(g["total_gerente"].sum())

    saldo = {cat: gerada[cat] - paga[cat] for cat, _, _ in VERBAS_CATS}
    for d in (gerada, paga, saldo):
        d["total"] = sum(d[cat] for cat, _, _ in VERBAS_CATS)
    return gerada, paga, saldo


def _status_mes(mes, pagos, tem_dado):
    """Rótulo do mês na tabela: só é 'Pago' quando consultor E gerente saíram."""
    if not tem_dado:
        return "—"
    flags = pagos.get(pd.Timestamp(mes), {})
    if flags.get("consultor") and flags.get("gerente"):
        return "Pago"
    if flags.get("consultor") or flags.get("gerente"):
        return "Parcial"
    return "Em aberto"


def pagina_verbas():
    st.title("💰 Verbas — Saldo e Pagamentos")
    # Dois erros diferentes, duas mensagens: tratar tudo como "falha de conexão"
    # já mascarou um deploy em que o app.py novo subiu sem o db.py novo.
    if not hasattr(db, "ler_vendas_verbas"):
        st.error("O `db.py` publicado está desatualizado: faltam as funções de "
                 "leitura da base de verbas. Suba a versão nova do `db.py` junto "
                 "com o `app.py` e reinicie o app.")
        return
    try:
        df = db.ler_vendas_verbas()
        pagos = db.ler_pagamentos_verba()
    except Exception as e:
        st.error("Não foi possível conectar ao banco de dados no momento. "
                 "Verifique se o MySQL está ativo e tente novamente.")
        st.caption(f"Detalhe técnico: {type(e).__name__}: {e}")
        return

    if df.empty:
        st.info("A base de verbas ainda não foi importada. Atualize a planilha "
                "*Base de dados para Dash Board.xlsx* e rode `python importar_verbas.py` "
                "na pasta do projeto.")
        return

    # ---- Filtro: ano + mês (o mês manda nos cards e no bloco de verbas) ----
    anos = sorted(df["data"].dt.year.unique(), reverse=True)
    cano, cmes, _ = st.columns([1, 2, 3])
    ano = cano.selectbox("Ano", anos, index=0)

    do_ano = df[df["data"].dt.year == ano]
    meses_disp = sorted(do_ano["data"].dt.to_period("M").unique(), reverse=True)
    labels = [OPCAO_ANO] + [mlabel(m.start_time) for m in meses_disp]
    mes_lbl = cmes.selectbox(
        "📅 Mês", labels, index=0,
        help="Filtra os cards e o bloco de verbas. A tabela e o gráfico abaixo "
             "sempre mostram os 12 meses do ano.")

    if mes_lbl == OPCAO_ANO:
        rec = do_ano
        periodo, anterior = f"{ano}", None
    else:
        p = meses_disp[labels.index(mes_lbl) - 1]
        rec = do_ano[do_ano["data"].dt.to_period("M") == p]
        periodo = mes_lbl
        # Mês anterior para o delta — pode não existir (primeiro mês da base).
        ant_p = p - 1
        prev = df[df["data"].dt.to_period("M") == ant_p]
        anterior = prev if not prev.empty else None

    ultima = df["data"].max()
    st.caption(f"Base atualizada até {ultima.strftime('%d/%m/%Y')} · "
               f"{len(df)} vendas · exibindo **{periodo}**")

    # ---- KPIs do período (tudo da mesma base) ----
    def kpis(d):
        pedidos = d["pedido"].nunique()
        return {
            "fat": float(d["total_item"].sum()),
            "diant": int(d.loc[d["tipo_refil"] == "diant", "qtde"].sum()),
            "tras": int(d.loc[d["tipo_refil"] == "tras", "qtde"].sum()),
            "pedidos": pedidos,
            "qtde": int(d["qtde"].sum()),
            "ticket": float(d["total_item"].sum()) / pedidos if pedidos else 0.0,
        }

    k = kpis(rec)
    ka = kpis(anterior) if anterior is not None else None

    def d(chave, kind):
        return delta_str(k[chave] - ka[chave], kind) if ka else None

    r1 = st.columns(3)
    r1[0].metric(f"Faturamento ({periodo})", fmt_money(k["fat"]), d("fat", "money"))
    r1[1].metric("Refil Diant. (pares)", fmt_int(k["diant"]), d("diant", "int"))
    r1[2].metric("Refil Tras. (un)", fmt_int(k["tras"]), d("tras", "int"))
    r2 = st.columns(3)
    r2[0].metric("Nº de Pedidos", fmt_int(k["pedidos"]), d("pedidos", "int"))
    r2[1].metric("Qtde total vendida", fmt_int(k["qtde"]), d("qtde", "int"))
    r2[2].metric("Ticket médio / pedido", fmt_money(k["ticket"]), d("ticket", "money"))
    if ka:
        st.caption("Variação comparada ao mês imediatamente anterior.")
    st.divider()

    # ---- Verba gerada / paga / saldo ----
    gerada, paga, saldo = _resumo_verbas(rec, pagos)

    st.subheader(f"Verba gerada — {periodo}")
    # key= gera a classe .st-key-verba_gerada, usada no CSS para reduzir a fonte
    # só desta linha (os valores aqui são os maiores da página).
    with st.container(key="verba_gerada"):
        cg = st.columns(4)
        for i, (cat, rotulo, _) in enumerate(VERBAS_CATS):
            cg[i].metric(rotulo, fmt_money(gerada[cat]))
        cg[3].metric("Total gerado", fmt_money(gerada["total"]))

    st.subheader("Verba já paga")
    cp = st.columns(4)
    cp[0].metric("Consultor", fmt_money(paga["consultor"]))
    cp[1].metric("Gerente", fmt_money(paga["gerente"]))
    cp[2].metric("Marketing", "—", help="A reserva de marketing não é paga a "
                                        "ninguém: entra inteira no saldo.")
    cp[3].metric("Total pago", fmt_money(paga["total"]))

    st.subheader("Saldo de verba")
    cs = st.columns(4)
    for i, (cat, rotulo, _) in enumerate(VERBAS_CATS):
        cs[i].metric(rotulo, fmt_money(saldo[cat]))
    cs[3].metric("SALDO TOTAL", fmt_money(saldo["total"]))

    with st.expander("Como o saldo é calculado"):
        em_aberto = []
        for p in sorted(rec["data"].dt.to_period("M").unique()):
            flags = pagos.get(pd.Timestamp(p.start_time), {})
            g = rec[rec["data"].dt.to_period("M") == p]
            partes = []
            if not flags.get("consultor"):
                partes.append(f"consultor {fmt_money(g['total_consultor'].sum())}")
            if not flags.get("gerente"):
                partes.append(f"gerente {fmt_money(g['total_gerente'].sum())}")
            partes.append(f"marketing {fmt_money(g['total_reserva'].sum())}")
            em_aberto.append(f"- **{mlabel(p.start_time)}**: " + " · ".join(partes))
        st.markdown(
            "**Saldo = verba gerada − verba paga**, mês a mês e por categoria.\n\n"
            "A verba de Marketing é paga separadamente para ser usada em eventos "
            "estratégicos de marketing. Consultor e gerente só saem do saldo nos "
            "meses marcados como pagos na aba **Pagamentos** da planilha.\n\n"
            "Composição do saldo no período exibido:\n\n" + "\n".join(em_aberto))
    st.divider()

    # ---- Tabela mês a mês (12 meses, independente do filtro) ----
    st.subheader(f"Verbas mês a mês — {ano}")
    por_mes = do_ano.groupby(do_ano["data"].dt.month)
    linhas, totais = [], dict(fat=0.0, c=0.0, g=0.0, m=0.0)
    for num in range(1, 13):
        tem = num in por_mes.groups
        gr = por_mes.get_group(num) if tem else None
        vals = {
            "fat": float(gr["total_item"].sum()) if tem else None,
            "c": float(gr["total_consultor"].sum()) if tem else None,
            "g": float(gr["total_gerente"].sum()) if tem else None,
            "m": float(gr["total_reserva"].sum()) if tem else None,
        }
        if tem:
            for chave in totais:
                totais[chave] += vals[chave]
        linhas.append({
            "Mês": MESES_PT[num],
            "Faturamento": fmt_money(vals["fat"]) if tem else "—",
            "Verba Consultor": fmt_money(vals["c"]) if tem else "—",
            "Verba Gerente": fmt_money(vals["g"]) if tem else "—",
            "Verba Marketing": fmt_money(vals["m"]) if tem else "—",
            "Total Verbas": fmt_money(vals["c"] + vals["g"] + vals["m"]) if tem else "—",
            "Status": _status_mes(dt.date(ano, num, 1), pagos, tem),
        })
    linhas.append({
        "Mês": "TOTAL",
        "Faturamento": fmt_money(totais["fat"]),
        "Verba Consultor": fmt_money(totais["c"]),
        "Verba Gerente": fmt_money(totais["g"]),
        "Verba Marketing": fmt_money(totais["m"]),
        "Total Verbas": fmt_money(totais["c"] + totais["g"] + totais["m"]),
        "Status": "",
    })
    st.dataframe(pd.DataFrame(linhas), width="stretch", hide_index=True)
    st.caption("“Pago” = verba de consultor E de gerente já quitadas naquele mês "
               "(marcação da aba Pagamentos). “Parcial” = só uma das duas.")

    # ---- Gráfico: 3 verbas mês a mês ----
    st.subheader("Evolução das verbas")
    dados = []
    for num in range(1, 13):
        gr = por_mes.get_group(num) if num in por_mes.groups else None
        for cat, rotulo, col in VERBAS_CATS:
            dados.append({"Mês": MESES_PT[num][:3],
                          "Verba": rotulo,
                          "Valor": float(gr[col].sum()) if gr is not None else 0.0})
    fig = px.bar(pd.DataFrame(dados), x="Mês", y="Valor", color="Verba",
                 barmode="group", color_discrete_sequence=PALETTE,
                 category_orders={"Mês": [m[:3] for m in MESES_PT[1:]],
                                  "Verba": [r for _, r, _ in VERBAS_CATS]})
    fig.update_traces(hovertemplate="%{x}<br>R$ %{y:,.2f}<extra>%{fullData.name}</extra>")
    fig.update_layout(template="plotly_dark", height=380,
                      margin=dict(l=10, r=10, t=30, b=10),
                      xaxis_title=None, yaxis_title=None,
                      paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                      font_color=BRANCO,
                      legend=dict(orientation="h", y=-0.15, title=None))
    fig.update_yaxes(rangemode="tozero", tickprefix="R$ ")
    st.plotly_chart(fig, width="stretch")


# ========================= ACESSO AO LANÇAMENTO =========================
def _senha_esperada():
    """Senha configurada: secrets tem prioridade, a constante é o padrão."""
    try:
        return str(st.secrets["acesso"]["senha_lancamento"])
    except (KeyError, FileNotFoundError):
        return SENHA_LANCAMENTO


def _senha_correta(senha):
    """Compara em tempo constante, em bytes (a senha pode ter acento)."""
    return bool(senha) and hmac.compare_digest(senha.encode("utf-8"),
                                               _senha_esperada().encode("utf-8"))


def _travar_lancamento():
    """Limpa o estado de acesso — o Lançamento volta a ficar oculto."""
    for chave in ("lanc_ok", "lanc_visto", "lanc_erros", "lanc_espera"):
        st.session_state.pop(chave, None)


def _revalidar_acesso():
    """Diz se o Lançamento está liberado e renova o relógio de inatividade.

    Tem efeito colateral: chamar uma única vez por execução, antes do menu.
    """
    if not st.session_state.get("lanc_ok"):
        return False
    agora = time.monotonic()
    if agora - st.session_state.get("lanc_visto", agora) > _TTL_LANCAMENTO:
        _travar_lancamento()
        return False
    st.session_state["lanc_visto"] = agora
    return True


@st.dialog("Área restrita")
def _dialogo_senha():
    """Pede a senha. Sem st.rerun(), o diálogo continua aberto com o erro."""
    st.caption("Informe a senha para liberar a aba de Lançamento.")
    with st.form("form_senha_lanc", clear_on_submit=True, border=False):
        senha = st.text_input("Senha", type="password",
                              label_visibility="collapsed", placeholder="Senha")
        enviou = st.form_submit_button("Entrar", type="primary", width="stretch")
    if not enviou:
        return

    restante = st.session_state.get("lanc_espera", 0.0) - time.monotonic()
    if restante > 0:
        st.error(f"Muitas tentativas. Aguarde {int(restante) + 1}s.")
        return

    if _senha_correta(senha):
        st.session_state["lanc_ok"] = True
        st.session_state["lanc_visto"] = time.monotonic()
        st.session_state["lanc_ir"] = True
        st.session_state.pop("lanc_erros", None)
        st.rerun()                       # fecha o diálogo e roda o app inteiro

    erros = st.session_state.get("lanc_erros", 0) + 1
    if erros >= _MAX_TENTATIVAS:
        st.session_state["lanc_espera"] = time.monotonic() + _ESPERA_BLOQUEIO
        st.session_state["lanc_erros"] = 0
        st.error("Muitas tentativas. Aguarde 1 minuto antes de tentar de novo.")
    else:
        st.session_state["lanc_erros"] = erros
        st.error(f"Senha incorreta. Tentativa {erros} de {_MAX_TENTATIVAS}.")


def _barra_acesso():
    """Cadeado discreto no fim da barra lateral. Chamar após montar a página."""
    st.sidebar.divider()
    liberado = bool(st.session_state.get("lanc_ok"))
    if st.sidebar.button("🔓" if liberado else "🔒", key="btn_cadeado",
                         type="tertiary",
                         help="Bloquear Lançamento" if liberado else "Área restrita"):
        if liberado:
            _travar_lancamento()
            st.rerun()
        else:
            _dialogo_senha()


# ============================ PÁGINA: LANÇAMENTO ============================
def meses_recentes(n=6):
    """Últimos n meses (1º dia de cada), do atual para trás."""
    hoje = dt.date.today().replace(day=1)
    y, m, out = hoje.year, hoje.month, []
    for _ in range(n):
        out.append(dt.date(y, m, 1))
        m -= 1
        if m == 0:
            m, y = 12, y - 1
    return out


def label_mes(d):
    """Rótulo do mês (ex.: 'Junho/2026')."""
    nomes = ["", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
             "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
    return f"{nomes[d.month]}/{d.year}"


def pagina_lancamento():
    # Trava redundante ao menu: garante que a página não renderize sem senha.
    if not st.session_state.get("lanc_ok"):
        st.error("Área restrita. Libere o acesso no cadeado da barra lateral.")
        return
    st.title("📝 Lançamento de dados")
    st.caption("Informe Passagens e Refis por consultor. O faturamento é calculado "
               "automaticamente a partir das quantidades.")

    unidades = db.listar_unidades()
    if not unidades:
        st.error("Nenhuma unidade cadastrada no banco.")
        return
    uni_por_nome = {u["nome_exibicao"]: u for u in unidades}
    nome_uni = st.selectbox("Unidade", list(uni_por_nome.keys()))
    uni = uni_por_nome[nome_uni]

    consultores = db.listar_consultores(uni["id"])
    if not consultores:
        st.warning("Esta unidade ainda não tem consultores cadastrados.")
        return
    cons_por_nome = {c["nome"]: c for c in consultores}
    nome_cons = st.selectbox("Consultor", list(cons_por_nome.keys()))
    cons = cons_por_nome[nome_cons]

    meses = meses_recentes(6)
    mes_por_label = {label_mes(d): d for d in meses}
    label_sel = st.selectbox("Mês de referência", list(mes_por_label.keys()))
    mes = mes_por_label[label_sel]

    existente = db.obter_lancamento(cons["id"], mes, uni["id"])
    if existente:
        st.info(f"Já existe lançamento para **{nome_cons}** em **{label_sel}**. "
                "Os valores abaixo estão preenchidos e serão atualizados ao salvar.")
    def_pass = int(existente["passagens"]) if existente and existente["passagens"] is not None else 0
    def_rd = int(existente["refil_diant"]) if existente else 0
    def_rt = int(existente["refil_tras"]) if existente else 0

    chave = f"{cons['id']}_{mes.isoformat()}"
    st.divider()
    c1, c2, c3 = st.columns(3)
    passagens = c1.number_input("Passagens", min_value=0, step=1, value=def_pass, key=f"pass_{chave}")
    refil_d = c2.number_input("Refil Dianteiro", min_value=0, step=1, value=def_rd, key=f"rd_{chave}")
    refil_t = c3.number_input("Refil Traseiro", min_value=0, step=1, value=def_rt, key=f"rt_{chave}")

    preco_d, preco_t = uni["preco_diant"], uni["preco_tras"]
    tot_d = refil_d * preco_d
    tot_t = refil_t * preco_t
    tot_g = tot_d + tot_t
    aprov = (refil_d / passagens) if passagens > 0 else None

    st.subheader("Prévia (calculada)")
    p1, p2, p3 = st.columns(3)
    p1.metric("Aproveitamento", fmt_pct(aprov))
    p2.metric("Faturamento dianteiro", fmt_money(tot_d))
    p3.metric("Faturamento total", fmt_money(tot_g))
    st.caption(f"Preços da marca {uni['marca']}: dianteiro {fmt_money(preco_d)} · "
               f"traseiro {fmt_money(preco_t)}.")

    if passagens > 0 and refil_d > passagens:
        st.warning("Refil Dianteiro maior que Passagens — o aproveitamento passou de 100%. Confira.")
    if passagens == 0 and (refil_d > 0 or refil_t > 0):
        st.warning("Passagens está zerado, mas há refis informados. Confira.")

    st.divider()
    if st.button("Salvar lançamento", type="primary", use_container_width=True):
        try:
            db.salvar_lancamento(cons["id"], mes, uni["id"], int(passagens), int(refil_d), int(refil_t))
        except Exception:
            st.error("Não foi possível salvar agora. Verifique a conexão com o banco e tente de novo.")
        else:
            st.success(f"Lançamento de **{nome_cons}** em **{label_sel}** salvo! "
                       f"Faturamento da semana: {fmt_money(tot_g)}.")
            st.toast("Dados gravados no banco.", icon="✅")

    # Excluir lançamento (só aparece quando já existe registro para o consultor/mês)
    if existente:
        with st.expander("Excluir este lançamento"):
            st.caption("Remove completamente o lançamento deste consultor nesta semana. "
                       "Use apenas se foi inserido por engano — não pode ser desfeito. "
                       "Para apenas corrigir um valor, basta editar acima e salvar.")
            ok = st.checkbox("Confirmo que quero excluir", key=f"conf_{chave}")
            if st.button("Excluir lançamento", disabled=not ok, key=f"del_{chave}"):
                try:
                    db.excluir_lancamento(cons["id"], mes, uni["id"])
                except Exception:
                    st.error("Não foi possível excluir agora. Verifique a conexão e tente de novo.")
                else:
                    st.success(f"Lançamento de **{nome_cons}** na semana **{label_sel}** excluído.")
                    st.toast("Lançamento removido.", icon="🗑️")


# ======================= PÁGINA: RELATÓRIO SEMANAL =======================
# Verba paga ao gerente da unidade por refil vendido (R$ por unidade).
VERBA_GER_DIANT, VERBA_GER_TRAS = 2.50, 1.25


def _relatorio_semana(df, mes_sel):
    """Monta o ranking por gerente/unidade de um mês + a linha de totais."""
    sem = df[df["mes"] == mes_sel].copy()
    sem["gerente"] = sem["gerente"].fillna("(sem gerente)")
    g = agg_by(sem, ["unidade", "gerente", "loja", "marca"])
    g["aprov_d"] = (g["refil_diant"] / g["passagens"]).where(g["passagens"] > 0)
    g["aprov_t"] = (g["refil_tras"] / g["passagens"]).where(g["passagens"] > 0)
    tot_fat = g["total_geral"].sum()
    g["part"] = (g["total_geral"] / tot_fat) if tot_fat else 0.0
    g["verba"] = g["refil_diant"] * VERBA_GER_DIANT + g["refil_tras"] * VERBA_GER_TRAS
    ov = agg_by(sem.assign(_g=1), "_g").iloc[0].to_dict()
    p = ov["passagens"]
    ov["aprov_d"] = (ov["refil_diant"] / p) if (p and p > 0) else None
    ov["aprov_t"] = (ov["refil_tras"] / p) if (p and p > 0) else None
    ov["verba"] = ov["refil_diant"] * VERBA_GER_DIANT + ov["refil_tras"] * VERBA_GER_TRAS
    return g, ov


def gerar_excel_semanal(g, ov, semana_lbl, ordenar):
    """Exporta o relatório semanal por gerente em Excel formatado."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    NAVY, WHITE = "FF1F3864", "FFFFFFFF"
    thin = Side(style="thin", color="FFD9D9D9")
    BD = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _v(x):
        return None if (isinstance(x, float) and pd.isna(x)) else x

    wb = Workbook(); ws = wb.active; ws.title = "Relatório Semanal"
    ws.cell(1, 1, f"Resultado DAHRUJ {semana_lbl}").font = \
        Font(name="Arial", bold=True, size=13, color=NAVY)
    ws.cell(2, 1, f"Ordenado por {ordenar} · Gerado em "
                  f"{dt.datetime.now().strftime('%d/%m/%Y %H:%M')}").font = \
        Font(name="Arial", size=9, color="FF666666")
    head = ["Seq", "Gerente", "Marca", "Loja", "Passagens", "Refil Diant.", "% Aprov",
            "Total Diant.", "Refil Tras.", "% Aprov", "Total Tras.", "Total Geral", "Part %",
            "Verba"]
    H = 4
    for j, h in enumerate(head, 1):
        c = ws.cell(H, j, h); c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
        c.fill = PatternFill("solid", fgColor=NAVY); c.border = BD
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, (_, r) in enumerate(g.iterrows()):
        rr = H + 1 + i
        vals = [i + 1, r["gerente"], r["marca"], r["loja"], r["passagens"], r["refil_diant"],
                r["aprov_d"], r["total_diant"], r["refil_tras"], r["aprov_t"],
                r["total_tras"], r["total_geral"], r["part"], r["verba"]]
        for j, v in enumerate(vals, 1):
            c = ws.cell(rr, j, _v(v)); c.font = Font(name="Arial", size=10); c.border = BD
            if j in (1, 5, 6, 7, 9, 10, 13):
                c.alignment = Alignment(horizontal="center")
            if j in (5, 6, 9):
                c.number_format = "#,##0"
            elif j in (7, 10, 13):
                c.number_format = "0.0%"
            elif j in (8, 11, 12, 14):
                c.number_format = "R$ #,##0.00"
        if i % 2 == 1:
            for j in range(1, len(head) + 1):
                ws.cell(rr, j).fill = PatternFill("solid", fgColor="FFF4F6FA")
    tr = H + 1 + len(g)
    ws.cell(tr, 2, "TOTAL").font = Font(name="Arial", bold=True, size=10)
    tvals = {5: _v(ov["passagens"]), 6: _v(ov["refil_diant"]), 7: _v(ov["aprov_d"]),
             8: _v(ov["total_diant"]), 9: _v(ov["refil_tras"]), 10: _v(ov["aprov_t"]),
             11: _v(ov["total_tras"]), 12: _v(ov["total_geral"]), 13: 1.0,
             14: _v(ov["verba"])}
    for j, v in tvals.items():
        c = ws.cell(tr, j, v); c.font = Font(name="Arial", bold=True, size=10); c.border = BD
        if j in (5, 6, 9):
            c.number_format = "#,##0"; c.alignment = Alignment(horizontal="center")
        elif j in (7, 10, 13):
            c.number_format = "0.0%"; c.alignment = Alignment(horizontal="center")
        elif j in (8, 11, 12, 14):
            c.number_format = "R$ #,##0.00"
    for j, w in enumerate([5, 16, 8, 16, 11, 12, 9, 13, 11, 9, 13, 14, 8, 12], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = f"A{H + 1}"
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def pagina_relatorio_semanal():
    st.title("📅 Relatório por Gerente")
    st.caption("Ranking das unidades no mês, no formato do relatório da diretoria.")
    try:
        df = db.ler_base_tidy()
    except Exception:
        st.error("Não foi possível conectar ao banco de dados no momento. "
                 "Verifique se o MySQL está ativo e tente novamente.")
        return
    if df.empty or df["mes"].dropna().empty:
        st.info("Ainda não há lançamentos no banco.")
        return

    meses = sorted(df["mes"].dropna().unique(), reverse=True)
    mes_por_label = {label_mes(pd.Timestamp(s).date()): s for s in meses}
    c1, c2 = st.columns([2, 1])
    label_sel = c1.selectbox("Mês", list(mes_por_label.keys()))
    ordenar = c2.radio("Ordenar por", ["Faturamento", "Aproveitamento"], horizontal=True)
    mes_sel = mes_por_label[label_sel]

    g, ov = _relatorio_semana(df, mes_sel)
    if g.empty:
        st.warning("Sem dados neste mês.")
        return
    ordcol = "total_geral" if ordenar == "Faturamento" else "aprov_d"
    g = g.sort_values(ordcol, ascending=False).reset_index(drop=True)

    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Passagens", fmt_int(ov["passagens"]))
    k2.metric("Refil Diant.", fmt_int(ov["refil_diant"]))
    k3.metric("Aproveitamento", fmt_pct(ov["aprov_d"]))
    k4.metric("Faturamento total", fmt_money(ov["total_geral"]))

    _, cexp = st.columns([3, 1])
    cexp.download_button(
        "📥 Exportar relatório (Excel)",
        data=gerar_excel_semanal(g, ov, label_sel, ordenar),
        file_name=f"Relatorio_Semanal_Dahruj_{pd.Timestamp(mes_sel).date().isoformat()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    linhas = []
    for i, (_, r) in enumerate(g.iterrows(), 1):
        linhas.append({
            "Seq": str(i), "Gerente": r["gerente"], "Marca": r["marca"], "Loja": r["loja"],
            "Passagens": fmt_int(r["passagens"]), "Refil Diant.": fmt_int(r["refil_diant"]),
            "% Aprov (D)": fmt_pct(r["aprov_d"]), "Total Diant.": fmt_money(r["total_diant"]),
            "Refil Tras.": fmt_int(r["refil_tras"]), "% Aprov (T)": fmt_pct(r["aprov_t"]),
            "Total Tras.": fmt_money(r["total_tras"]), "Total Geral": fmt_money(r["total_geral"]),
            "Part %": fmt_pct(r["part"]), "Verba": fmt_money(r["verba"]),
        })
    linhas.append({
        "Seq": "", "Gerente": "TOTAL", "Marca": "", "Loja": "",
        "Passagens": fmt_int(ov["passagens"]), "Refil Diant.": fmt_int(ov["refil_diant"]),
        "% Aprov (D)": fmt_pct(ov["aprov_d"]), "Total Diant.": fmt_money(ov["total_diant"]),
        "Refil Tras.": fmt_int(ov["refil_tras"]), "% Aprov (T)": fmt_pct(ov["aprov_t"]),
        "Total Tras.": fmt_money(ov["total_tras"]), "Total Geral": fmt_money(ov["total_geral"]),
        "Part %": "100,0%", "Verba": fmt_money(ov["verba"]),
    })
    st.dataframe(pd.DataFrame(linhas), use_container_width=True, hide_index=True)
    st.caption(f"Verba = (Refil Diant. × R$ {fmt_money(VERBA_GER_DIANT)[3:]}) + "
               f"(Refil Tras. × R$ {fmt_money(VERBA_GER_TRAS)[3:]}). "
               "Part % = participação da unidade no faturamento total da semana.")


# ===================== PÁGINA: RELATÓRIO POR CONSULTOR =====================
# Verba paga ao consultor por refil vendido (R$ por unidade).
VERBA_DIANT, VERBA_TRAS = 10, 5


def _relatorio_consultor(df, mes_sel):
    """Monta o ranking por consultor/unidade de um mês + a linha de totais."""
    sem = df[df["mes"] == mes_sel].copy()
    g = agg_by(sem, ["consultor", "unidade"])
    g["aprov_d"] = (g["refil_diant"] / g["passagens"]).where(g["passagens"] > 0)
    g["aprov_t"] = (g["refil_tras"] / g["passagens"]).where(g["passagens"] > 0)
    tot_fat = g["total_geral"].sum()
    g["part"] = (g["total_geral"] / tot_fat) if tot_fat else 0.0
    g["verba"] = g["refil_diant"] * VERBA_DIANT + g["refil_tras"] * VERBA_TRAS
    ov = agg_by(sem.assign(_g=1), "_g").iloc[0].to_dict()
    p = ov["passagens"]
    ov["aprov_d"] = (ov["refil_diant"] / p) if (p and p > 0) else None
    ov["aprov_t"] = (ov["refil_tras"] / p) if (p and p > 0) else None
    ov["verba"] = ov["refil_diant"] * VERBA_DIANT + ov["refil_tras"] * VERBA_TRAS
    return g, ov


def gerar_excel_consultor(g, ov, mes_lbl, ordenar):
    """Exporta o relatório mensal por consultor em Excel formatado."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    NAVY, WHITE = "FF1F3864", "FFFFFFFF"
    thin = Side(style="thin", color="FFD9D9D9")
    BD = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _v(x):
        return None if (isinstance(x, float) and pd.isna(x)) else x

    wb = Workbook(); ws = wb.active; ws.title = "Relatório Consultor"
    ws.cell(1, 1, f"Resultado por Consultor DAHRUJ {mes_lbl}").font = \
        Font(name="Arial", bold=True, size=13, color=NAVY)
    ws.cell(2, 1, f"Ordenado por {ordenar} · Gerado em "
                  f"{dt.datetime.now().strftime('%d/%m/%Y %H:%M')}").font = \
        Font(name="Arial", size=9, color="FF666666")
    head = ["Seq", "Consultor", "Unidade", "Passagens", "Refil Diant.", "% Aprov",
            "Total Diant.", "Refil Tras.", "% Aprov", "Total Tras.", "Total Geral",
            "Part %", "Verba"]
    COL_INT, COL_PCT, COL_MONEY = (4, 5, 8), (6, 9, 12), (7, 10, 11, 13)
    H = 4
    for j, h in enumerate(head, 1):
        c = ws.cell(H, j, h); c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
        c.fill = PatternFill("solid", fgColor=NAVY); c.border = BD
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, (_, r) in enumerate(g.iterrows()):
        rr = H + 1 + i
        vals = [i + 1, r["consultor"], r["unidade"], r["passagens"], r["refil_diant"],
                r["aprov_d"], r["total_diant"], r["refil_tras"], r["aprov_t"],
                r["total_tras"], r["total_geral"], r["part"], r["verba"]]
        for j, v in enumerate(vals, 1):
            c = ws.cell(rr, j, _v(v)); c.font = Font(name="Arial", size=10); c.border = BD
            if j == 1 or j in COL_INT or j in COL_PCT:
                c.alignment = Alignment(horizontal="center")
            if j in COL_INT:
                c.number_format = "#,##0"
            elif j in COL_PCT:
                c.number_format = "0.0%"
            elif j in COL_MONEY:
                c.number_format = "R$ #,##0.00"
        if i % 2 == 1:
            for j in range(1, len(head) + 1):
                ws.cell(rr, j).fill = PatternFill("solid", fgColor="FFF4F6FA")
    tr = H + 1 + len(g)
    ws.cell(tr, 2, "TOTAL").font = Font(name="Arial", bold=True, size=10)
    tvals = {4: _v(ov["passagens"]), 5: _v(ov["refil_diant"]), 6: _v(ov["aprov_d"]),
             7: _v(ov["total_diant"]), 8: _v(ov["refil_tras"]), 9: _v(ov["aprov_t"]),
             10: _v(ov["total_tras"]), 11: _v(ov["total_geral"]), 12: 1.0,
             13: _v(ov["verba"])}
    for j, v in tvals.items():
        c = ws.cell(tr, j, v); c.font = Font(name="Arial", bold=True, size=10); c.border = BD
        if j in COL_INT:
            c.number_format = "#,##0"; c.alignment = Alignment(horizontal="center")
        elif j in COL_PCT:
            c.number_format = "0.0%"; c.alignment = Alignment(horizontal="center")
        elif j in COL_MONEY:
            c.number_format = "R$ #,##0.00"
    for j, w in enumerate([5, 22, 20, 11, 12, 9, 13, 11, 9, 13, 14, 8, 12], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = f"A{H + 1}"
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def pagina_relatorio_consultor():
    st.title("🧑‍💼 Relatório por Consultor")
    st.caption("Ranking dos consultores no mês, com a verba de cada um.")
    try:
        df = db.ler_base_tidy()
    except Exception:
        st.error("Não foi possível conectar ao banco de dados no momento. "
                 "Verifique se o MySQL está ativo e tente novamente.")
        return
    if df.empty or df["mes"].dropna().empty:
        st.info("Ainda não há lançamentos no banco.")
        return

    meses = sorted(df["mes"].dropna().unique(), reverse=True)
    mes_por_label = {label_mes(pd.Timestamp(s).date()): s for s in meses}
    TODAS = "Todas as unidades"
    unidades = [TODAS] + sorted(df["unidade"].dropna().unique())
    c1, c2, c3 = st.columns([2, 2, 1])
    label_sel = c1.selectbox("Mês", list(mes_por_label.keys()))
    unidade_sel = c2.selectbox("Unidade", unidades)
    ordenar = c3.radio("Ordenar por", ["Faturamento", "Aproveitamento"], horizontal=True)
    mes_sel = mes_por_label[label_sel]

    if unidade_sel != TODAS:
        df = df[df["unidade"] == unidade_sel]
    if df[df["mes"] == mes_sel].empty:
        st.warning("Sem dados para esta unidade neste mês.")
        return
    g, ov = _relatorio_consultor(df, mes_sel)
    if g.empty:
        st.warning("Sem dados neste mês.")
        return
    ordcol = "total_geral" if ordenar == "Faturamento" else "aprov_d"
    g = g.sort_values(ordcol, ascending=False).reset_index(drop=True)

    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Passagens", fmt_int(ov["passagens"]))
    k2.metric("Refil Diant.", fmt_int(ov["refil_diant"]))
    k3.metric("Aproveitamento", fmt_pct(ov["aprov_d"]))
    k4.metric("Faturamento total", fmt_money(ov["total_geral"]))

    _, cexp = st.columns([3, 1])
    cexp.download_button(
        "📥 Exportar relatório (Excel)",
        data=gerar_excel_consultor(g, ov, label_sel, ordenar),
        file_name=f"Relatorio_Consultor_Dahruj_{pd.Timestamp(mes_sel).date().isoformat()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    linhas = []
    for i, (_, r) in enumerate(g.iterrows(), 1):
        linhas.append({
            "Seq": str(i), "Consultor": r["consultor"], "Unidade": r["unidade"],
            "Passagens": fmt_int(r["passagens"]), "Refil Diant.": fmt_int(r["refil_diant"]),
            "% Aprov (D)": fmt_pct(r["aprov_d"]), "Total Diant.": fmt_money(r["total_diant"]),
            "Refil Tras.": fmt_int(r["refil_tras"]), "% Aprov (T)": fmt_pct(r["aprov_t"]),
            "Total Tras.": fmt_money(r["total_tras"]), "Total Geral": fmt_money(r["total_geral"]),
            "Part %": fmt_pct(r["part"]), "Verba": fmt_money(r["verba"]),
        })
    linhas.append({
        "Seq": "", "Consultor": "TOTAL", "Unidade": "",
        "Passagens": fmt_int(ov["passagens"]), "Refil Diant.": fmt_int(ov["refil_diant"]),
        "% Aprov (D)": fmt_pct(ov["aprov_d"]), "Total Diant.": fmt_money(ov["total_diant"]),
        "Refil Tras.": fmt_int(ov["refil_tras"]), "% Aprov (T)": fmt_pct(ov["aprov_t"]),
        "Total Tras.": fmt_money(ov["total_tras"]), "Total Geral": fmt_money(ov["total_geral"]),
        "Part %": "100,0%", "Verba": fmt_money(ov["verba"]),
    })
    st.dataframe(pd.DataFrame(linhas), use_container_width=True, hide_index=True)
    st.caption(f"Verba = (Refil Diant. × R$ {VERBA_DIANT}) + (Refil Tras. × R$ {VERBA_TRAS}). "
               "Part % = participação do consultor no faturamento total do mês.")


# ============================ NAVEGAÇÃO ============================
_injetar_css()
mostrar_logo()
st.sidebar.title("Dahruj")

# Normalização ANTES do menu: com o widget key="menu_pagina" já criado, escrever
# em st.session_state["menu_pagina"] levanta StreamlitAPIException.
_ir_para_lanc = st.session_state.pop("lanc_ir", False)
_liberado = _revalidar_acesso()
if _liberado and _ir_para_lanc:
    st.session_state["menu_pagina"] = PAG_LANCAMENTO
elif not _liberado and st.session_state.get("menu_pagina") == PAG_LANCAMENTO:
    st.session_state["menu_pagina"] = PAG_DASHBOARD

opcoes = PAGINAS_PUBLICAS + ([PAG_LANCAMENTO] if _liberado else [])
pagina = st.sidebar.radio("Menu", opcoes, key="menu_pagina")
st.sidebar.divider()

DESTINOS = {
    PAG_DASHBOARD: pagina_dashboard,
    PAG_VERBAS: pagina_verbas,
    PAG_GERENTE: pagina_relatorio_semanal,
    PAG_CONSULTOR: pagina_relatorio_consultor,
    PAG_LANCAMENTO: pagina_lancamento,
}
try:
    DESTINOS.get(pagina, pagina_dashboard)()
finally:
    _barra_acesso()                      # cadeado sobrevive a erro na página